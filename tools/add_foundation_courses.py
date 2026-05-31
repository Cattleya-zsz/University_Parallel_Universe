from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


path = r"D:\Uni_Parallel\数据统计例表.xlsx"
wb = load_workbook(path)
ws = wb["核心课程"]

headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
header_index = {name: idx + 1 for idx, name in enumerate(headers)}

required_headers = [
    "major",
    "course_category",
    "course_group",
    "course_name",
    "brief_intro",
    "usage_hint",
    "source",
    "source_url",
]

if headers != required_headers:
    raise ValueError(f"Unexpected headers: {headers}")

existing = set()
for row in range(2, ws.max_row + 1):
    major = ws.cell(row, header_index["major"]).value
    course_name = ws.cell(row, header_index["course_name"]).value
    existing.add((major, course_name))

rows = [
    [
        "计算机类",
        "专业基础课",
        "数学基础",
        "高等数学",
        "学习函数、极限、微积分等内容，是后续算法分析和数学建模的重要基础。",
        "适合生成数学课、习题课、课后刷题等上课事件。",
        "南方科技大学计算机科学与技术专业介绍",
        "https://cse.sustech.edu.cn/undergraduate/",
    ],
    [
        "计算机类",
        "专业基础课",
        "数学基础",
        "线性代数",
        "学习向量、矩阵和线性变换，是图形学、机器学习和数据分析常用基础。",
        "适合生成基础数学课、推导练习、AI方向准备等事件。",
        "南方科技大学计算机科学与技术专业介绍",
        "https://cse.sustech.edu.cn/undergraduate/",
    ],
    [
        "计算机类",
        "专业基础课",
        "数学基础",
        "概率论与数理统计",
        "学习随机现象、概率分布和统计推断，是机器学习、数据挖掘等方向的重要基础。",
        "适合生成概率统计课、数据分析准备、考前复习等事件。",
        "南方科技大学计算机科学与技术专业介绍",
        "https://cse.sustech.edu.cn/undergraduate/",
    ],
    [
        "计算机类",
        "专业基础课",
        "工程基础",
        "程序设计基础",
        "学习编程语法、控制结构和基本问题求解方法，是进入专业课前的关键起点。",
        "适合生成新生编程课、上机练习、基础作业等事件。",
        "南方科技大学计算机科学与技术专业介绍",
        "https://cse.sustech.edu.cn/undergraduate/",
    ],
    [
        "医学类",
        "专业基础课",
        "自然科学基础",
        "医用化学",
        "学习医学相关的基础化学知识，为生物化学、药理学等课程打底。",
        "适合生成基础课、实验准备、前期预习等事件。",
        "延安大学延安医学院专业介绍",
        "https://yxy.yau.edu.cn/bkjy/zyjs.htm",
    ],
    [
        "医学类",
        "专业基础课",
        "自然科学基础",
        "医用物理学",
        "学习医学成像、力学和电学等相关物理基础，帮助理解仪器与人体机制。",
        "适合生成基础理论课、实验导论、影像理解等事件。",
        "延安大学延安医学院专业介绍",
        "https://yxy.yau.edu.cn/bkjy/zyjs.htm",
    ],
    [
        "医学类",
        "专业基础课",
        "生物学基础",
        "细胞生物学",
        "学习细胞结构、功能和生命活动机制，是后续基础医学课程的重要前置知识。",
        "适合生成基础生物课、显微观察、背诵复习等事件。",
        "延安大学延安医学院专业介绍",
        "https://yxy.yau.edu.cn/bkjy/zyjs.htm",
    ],
    [
        "医学类",
        "专业基础课",
        "生物学基础",
        "医学遗传学",
        "学习遗传规律、基因与疾病之间的关系，为临床和科研理解提供基础。",
        "适合生成理论课、病例分析、拓展阅读等事件。",
        "延安大学延安医学院专业介绍",
        "https://yxy.yau.edu.cn/bkjy/zyjs.htm",
    ],
    [
        "经管类",
        "专业基础课",
        "数学与方法",
        "高等数学",
        "学习微积分等基本数学工具，为经济学分析、统计学和计量方法提供基础。",
        "适合生成基础数学课、习题训练、考试复习等事件。",
        "东北财经大学经济学专业培养方案",
        "https://econ.dufe.edu.cn/content_90636.html",
    ],
    [
        "经管类",
        "专业基础课",
        "数学与方法",
        "线性代数",
        "学习矩阵和线性模型，是统计分析、计量经济学和数据建模的重要工具。",
        "适合生成数学基础课、模型理解、数据方法预备等事件。",
        "东北财经大学经济学专业培养方案",
        "https://econ.dufe.edu.cn/content_90636.html",
    ],
    [
        "经管类",
        "专业基础课",
        "数学与方法",
        "概率论与数理统计",
        "学习概率模型和统计推断，是数据分析、市场研究和计量方法的常用基础。",
        "适合生成统计基础课、问卷分析、商业数据准备等事件。",
        "东北财经大学经济学专业培养方案",
        "https://econ.dufe.edu.cn/content_90636.html",
    ],
    [
        "经管类",
        "专业基础课",
        "工具基础",
        "经济法基础",
        "学习企业经营和经济活动中常见的法律规则，帮助理解商业运行边界。",
        "适合生成基础法商课、案例讨论、考试复习等事件。",
        "中山大学工商管理专业介绍",
        "https://isbf.sysu.edu.cn/zh-hans/article/329",
    ],
]

added = 0
for row in rows:
    key = (row[0], row[3])
    if key in existing:
        continue
    ws.append(row)
    added += 1

header_fill = PatternFill("solid", fgColor="1F4E78")
white_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

widths = [14, 14, 16, 20, 42, 42, 36, 48]
for i, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

wb.save(path)
print(f"added {added} foundation courses; total rows now {ws.max_row}")
