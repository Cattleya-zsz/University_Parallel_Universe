from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


path = r"D:\Uni_Parallel\数据统计例表.xlsx"
wb = load_workbook(path)
ws = wb["核心课程"]

if ws.cell(1, ws.max_column).value != "source_url":
    ws.cell(1, ws.max_column + 1, "source_url")

urls = {
    "南方科技大学计算机科学与技术专业介绍": "https://cse.sustech.edu.cn/undergraduate/",
    "北京师范大学-香港浸会大学联合国际学院课程设置": "https://www.uic.edu.cn/virtual_attach_file.vsb?afc=nL8LS8o7CPozliozzMkMmviL7LaLR7lin7UbMzlPMlQRUNU0gihFp2hmCIa0n1yPoSh7n1y8oz9snmQ7LlrfMm6kMzrRMmvDLlLiM4CYM4MFUlVknRvsolQFMlCbLm-Jv2nto4OeosrXCDMJgDTJQty0Lz7YLkyPLzGZokbw62r8c&e=.pdf&nid=11088&tid=1120",
    "延安大学延安医学院专业介绍": "https://yxy.yau.edu.cn/bkjy/zyjs.htm",
    "东北财经大学经济学专业培养方案；中山大学工商管理专业介绍": "https://econ.dufe.edu.cn/content_90636.html ; https://isbf.sysu.edu.cn/zh-hans/article/329",
    "中山大学工商管理专业介绍": "https://isbf.sysu.edu.cn/zh-hans/article/329",
    "中山大学工商管理专业介绍；东北财经大学经济学专业培养方案": "https://isbf.sysu.edu.cn/zh-hans/article/329 ; https://econ.dufe.edu.cn/content_90636.html",
    "东北财经大学经济学专业培养方案": "https://econ.dufe.edu.cn/content_90636.html",
}

for row in range(2, ws.max_row + 1):
    source = ws.cell(row, 7).value
    ws.cell(row, 8, urls.get(source, ""))

header_fill = PatternFill("solid", fgColor="1F4E78")
white_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions[get_column_letter(8)].width = 48
ws.auto_filter.ref = f"A1:H{ws.max_row}"

wb.save(path)
print(f"updated source urls: {ws.max_row} rows, {ws.max_column} columns")
