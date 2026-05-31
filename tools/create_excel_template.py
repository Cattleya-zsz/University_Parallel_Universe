from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


out = Path(r"D:\Uni_Parallel\social_samples_template.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "数据记录"

headers = [
    "sample_id",
    "major",
    "activity_period",
    "activity_order",
    "activity_type",
    "activity_description",
    "location_type",
    "notes",
]
header_notes = [
    "样本编号，同一条vlog多行共用一个编号，如CS_001",
    "专业分类，使用下拉选项",
    "活动发生时间段，使用下拉选项",
    "活动顺序，填1、2、3...",
    "标准化活动类型，使用下拉选项",
    "用一句话描述该活动，不写隐私信息",
    "标准化地点类型，使用下拉选项",
    "仅写非隐私备注，如普通日/考试周/周末",
]

ws.append(headers)
ws.append(header_notes)

samples = [
    ["CS_001", "计算机类", "早晨", 1, "通勤/移动", "从宿舍出发前往教学楼", "宿舍", "普通日样本"],
    ["CS_001", "计算机类", "上午", 2, "上课", "专业基础课或理论课", "教学楼", "普通日样本"],
    ["CS_001", "计算机类", "下午", 3, "项目实践", "写代码或完成课程项目", "实验楼", "普通日样本"],
    ["CS_001", "计算机类", "晚上", 4, "自习", "复习课程内容或刷题", "图书馆", "普通日样本"],
    ["MED_001", "医学类", "上午", 1, "上课", "医学基础课程学习", "教学楼", "普通日样本"],
    ["MED_001", "医学类", "下午", 2, "实验", "实验课或技能训练", "实验楼", "普通日样本"],
    ["MED_001", "医学类", "晚上", 3, "自习", "背诵和复习专业内容", "图书馆", "考试周样本"],
    ["BUS_001", "经管类", "上午", 1, "上课", "管理或经济类课程", "教学楼", "普通日样本"],
    ["BUS_001", "经管类", "下午", 2, "小组讨论", "准备课程展示或案例分析", "自习室", "普通日样本"],
    ["BUS_001", "经管类", "晚上", 3, "社团活动", "参加学生组织或社团活动", "社团活动中心", "普通日样本"],
]
for row in samples:
    ws.append(row)

header_fill = PatternFill("solid", fgColor="1F4E78")
note_fill = PatternFill("solid", fgColor="D9EAF7")
white_font = Font(color="FFFFFF", bold=True)
small_font = Font(size=9, color="44546A")
thin = Side(style="thin", color="D9E2F3")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
for cell in ws[2]:
    cell.fill = note_fill
    cell.font = small_font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

widths = [14, 14, 14, 12, 16, 34, 16, 22]
for i, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 45
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{ws.max_row}"

lists = wb.create_sheet("下拉选项")
majors = ["计算机类", "医学类", "经管类", "法学类", "建筑/设计类"]
periods = ["早晨", "上午", "中午", "下午", "傍晚", "晚上", "睡前"]
activities = [
    "上课",
    "自习",
    "实验",
    "项目实践",
    "小组讨论",
    "吃饭",
    "运动",
    "社团活动",
    "兼职/实习",
    "娱乐放松",
    "休息",
    "通勤/移动",
    "其他",
]
locations = [
    "宿舍",
    "教学楼",
    "实验楼",
    "图书馆",
    "自习室",
    "食堂",
    "操场",
    "体育馆",
    "社团活动中心",
    "校外",
    "线上",
    "未知",
]
columns = [
    ("major", majors),
    ("activity_period", periods),
    ("activity_type", activities),
    ("location_type", locations),
]
for col_idx, (title, values) in enumerate(columns, start=1):
    lists.cell(row=1, column=col_idx, value=title)
    lists.cell(row=1, column=col_idx).fill = header_fill
    lists.cell(row=1, column=col_idx).font = white_font
    for row_idx, value in enumerate(values, start=2):
        lists.cell(row=row_idx, column=col_idx, value=value)
    lists.column_dimensions[get_column_letter(col_idx)].width = 18

validations = [
    ("B3:B502", "=下拉选项!$A$2:$A$6"),
    ("C3:C502", "=下拉选项!$B$2:$B$8"),
    ("E3:E502", "=下拉选项!$C$2:$C$14"),
    ("G3:G502", "=下拉选项!$D$2:$D$13"),
]
for cell_range, formula in validations:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "请选择下拉列表中的标准值，避免后续清洗数据。"
    dv.errorTitle = "非标准值"
    ws.add_data_validation(dv)
    dv.add(cell_range)

stats = wb.create_sheet("统计概览")
stats["A1"] = "专业样本数统计"
stats["A1"].font = Font(bold=True, size=14)
stats.append(["专业", "活动记录数"])
for idx, major in enumerate(majors, start=3):
    stats.cell(row=idx, column=1, value=major)
    stats.cell(row=idx, column=2, value=f'=COUNTIF(数据记录!$B:$B,A{idx})')

stats["D1"] = "活动类型统计"
stats["D1"].font = Font(bold=True, size=14)
stats["D2"] = "活动类型"
stats["E2"] = "记录数"
for idx, act in enumerate(activities, start=3):
    stats.cell(row=idx, column=4, value=act)
    stats.cell(row=idx, column=5, value=f'=COUNTIF(数据记录!$E:$E,D{idx})')

for cell in stats[2]:
    if cell.value:
        cell.fill = header_fill
        cell.font = white_font
for col in ["A", "B", "D", "E"]:
    stats.column_dimensions[col].width = 18

chart = BarChart()
chart.title = "各专业活动记录数"
chart.y_axis.title = "记录数"
chart.x_axis.title = "专业"
data = Reference(stats, min_col=2, min_row=2, max_row=2 + len(majors))
cats = Reference(stats, min_col=1, min_row=3, max_row=2 + len(majors))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 7
chart.width = 12
stats.add_chart(chart, "G2")

readme = wb.create_sheet("填写说明")
content = [
    ["填写目标", "只记录公开vlog中的日程安排信息，不记录隐私敏感信息。"],
    ["一条vlog如何记录", "同一条vlog使用同一个sample_id，每出现一个活动就新增一行。"],
    ["sample_id规则", "计算机类CS_001，医学类MED_001，经管类BUS_001，法学类LAW_001，建筑/设计类DES_001。"],
    ["activity_order", "按当天出现顺序填写1、2、3、4。"],
    ["activity_description", "一句话描述活动即可，例如“专业课”“图书馆复习”“小组案例讨论”。"],
    ["notes", "只能写普通日、考试周、周末等非隐私备注。"],
    ["不要记录", "博主昵称、头像、真实姓名、具体学校、宿舍楼号、视频链接、评论区个人信息、可识别截图。"],
    ["后续处理", "导出CSV后交给AI，归纳各专业典型事件、选项和评分逻辑。"],
]
readme.append(["项目", "说明"])
for row in content:
    readme.append(row)
for cell in readme[1]:
    cell.fill = header_fill
    cell.font = white_font
for col in ["A", "B"]:
    readme.column_dimensions[col].width = 24 if col == "A" else 80
for row in readme.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

wb.save(out)
print(out)
