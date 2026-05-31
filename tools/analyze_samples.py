from collections import Counter, defaultdict
from openpyxl import load_workbook


path = r"D:\Uni_Parallel\数据统计例表.xlsx"
wb = load_workbook(path, data_only=True)
ws = wb["数据记录"]
headers = [cell.value for cell in ws[1]]

rows = []
for raw in ws.iter_rows(min_row=3, values_only=True):
    if any(value is not None for value in raw):
        rows.append(dict(zip(headers, raw)))

print("headers:", headers)
print("records:", len(rows))

by_major = defaultdict(set)
records_by_major = Counter()
period = defaultdict(Counter)
activity = defaultdict(Counter)
location = defaultdict(Counter)
missing = Counter()
duplicate_orders = defaultdict(list)
sample_orders = defaultdict(list)

for row in rows:
    major = row.get("major")
    sample_id = row.get("sample_id")
    if major and sample_id:
        by_major[major].add(sample_id)
        sample_orders[(major, sample_id)].append(row.get("activity_order"))
    records_by_major[major] += 1
    period[major][row.get("activity_period")] += 1
    activity[major][row.get("activity_type")] += 1
    location[major][row.get("location_type")] += 1
    for header in headers:
        if row.get(header) in (None, ""):
            missing[header] += 1

for key, orders in sample_orders.items():
    clean = [order for order in orders if order is not None]
    if len(clean) != len(set(clean)):
        duplicate_orders[key].append(clean)

for major in sorted(by_major):
    print(f"\nMAJOR: {major}")
    print("unique_samples:", len(by_major[major]))
    print("activity_records:", records_by_major[major])
    print("periods:", dict(period[major]))
    print("activities:", dict(activity[major]))
    print("locations:", dict(location[major]))

print("\nmissing:", dict(missing))
print("duplicate_orders:", dict(duplicate_orders))

