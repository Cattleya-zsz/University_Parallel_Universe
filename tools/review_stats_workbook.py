from collections import Counter, defaultdict

from openpyxl import load_workbook


path = r"D:\Uni_Parallel\数据统计例表.xlsx"
wb = load_workbook(path, data_only=True)


def normalize(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def review_data_sheet():
    ws = wb["数据记录"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for row_idx in range(3, ws.max_row + 1):
        row = {headers[col - 1]: normalize(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)}
        if any(value is not None for value in row.values()):
            row["_row"] = row_idx
            rows.append(row)

    by_major_samples = defaultdict(set)
    by_major_records = Counter()
    by_major_periods = defaultdict(Counter)
    by_major_activities = defaultdict(Counter)
    by_major_locations = defaultdict(Counter)
    missing = Counter()
    incomplete_rows = []

    for row in rows:
        major = row["major"]
        sample_id = row["sample_id"]
        if major and sample_id:
            by_major_samples[major].add(sample_id)
        by_major_records[major] += 1
        by_major_periods[major][row["activity_period"]] += 1
        by_major_activities[major][row["activity_type"]] += 1
        by_major_locations[major][row["location_type"]] += 1

        row_missing = []
        for key in headers:
            if row[key] is None:
                missing[key] += 1
                row_missing.append(key)
        if row_missing:
            incomplete_rows.append((row["_row"], row_missing, row))

    print("=== 数据记录 ===")
    print("max_row:", ws.max_row)
    print("effective_rows:", len(rows))
    for major in sorted(by_major_records, key=lambda item: "" if item is None else str(item)):
        print(f"\n[{major}]")
        print("samples:", len(by_major_samples[major]))
        print("records:", by_major_records[major])
        print("periods:", dict(by_major_periods[major]))
        print("activities:", dict(by_major_activities[major]))
        print("locations:", dict(by_major_locations[major]))

    print("\nmissing_counts:", dict(missing))
    print("incomplete_row_count:", len(incomplete_rows))
    print("first_incomplete_rows:")
    for item in incomplete_rows[:10]:
        print(item)


def review_course_sheet():
    ws = wb["核心课程"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {headers[col - 1]: normalize(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)}
        if any(value is not None for value in row.values()):
            rows.append(row)

    by_major = defaultdict(Counter)
    missing = Counter()
    for row in rows:
        by_major[row["major"]][row["course_category"]] += 1
        for key in headers:
            if row[key] is None:
                missing[key] += 1

    print("\n=== 核心课程 ===")
    print("max_row:", ws.max_row)
    print("effective_rows:", len(rows))
    print("headers:", headers)
    print("by_major:", {major: dict(counter) for major, counter in by_major.items()})
    print("missing_counts:", dict(missing))


review_data_sheet()
review_course_sheet()
